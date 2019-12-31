import requests
from time import gmtime, strftime
import time
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import urllib3
import io
urllib3.disable_warnings()


def utc2local(utc_st):
    """UTC时间转本地时间（+8: 00）"""
    now_stamp = time.time()
    local_time = datetime.datetime.fromtimestamp(now_stamp)
    utc_time = datetime.datetime.utcfromtimestamp(now_stamp)
    offset = local_time - utc_time
    local_st = utc_st + offset
    return local_st

def utcstr2localstr(utc_time):
    """utc_time format is %Y-%m-%d %H:%M:%S"""
    format = '%Y-%m-%d %H:%M:%S'
    utcdatetime = datetime.datetime.strptime(utc_time, format)
    localdatetime = utc2local(utcdatetime)
    # if outformat:
    #     format = outformat

    format = '%m-%d %H:%M'
    return localdatetime.strftime(format)

def get(url):
    return requests.get(url)

def get_all_leagues():
    url = 'https://api.dongqiudi.com/v2/config/match_menu?mark=gif&platform=iphone&version=731&home_id=50000515'
    r = get(url)
    data = r.json()
    result = []
    for one in data['data']['list']:
        if one['type'] == 'league':
            result.append(one)
    return result

def get_all_matches(match_url):
    # https://api.dongqiudi.com/data/tab/league/new/4
    url = match_url
    current_time = strftime("%Y-%m-%d%H:%M:%S", gmtime())
    url = '%s?%s&init=1' % (url, current_time)
    r = get(url)
    data = r.json()['list']
    return data

def get_team_info(teamid):
    url = 'https://api.dongqiudi.com/data/v1/sample/team/%s' % teamid
    r = get(url)
    return r.json()

def get_pre_analyze_data_contrast(matchid):
    # https://sport-data.dongqiudi.com/soccer/biz/dqd/match/pre_analyze_data_contrast/52892596?app=dqd
    url = 'https://sport-data.dongqiudi.com/soccer/biz/dqd/match/pre_analyze_data_contrast/%s?app=dqd' % matchid
    r = get(url)
    data = r.json()['data']
    return data

def get_main_person(teamid):
    # https://sport-data.dongqiudi.com/soccer/biz/dqd/team/statistic/50000516?lang=zh-cn
    url = 'https://sport-data.dongqiudi.com/soccer/biz/dqd/team/statistic/%s?lang=zh-cn' % teamid
    r = get(url)
    return r.json()

def getProperty(data, key, value):
    for one in data:
        if one[key] == value:
            return one

def downloadImg(url):
    http = urllib3.PoolManager()
    r = http.request('GET', url)
    image_file = io.BytesIO(r.data)
    img = Image(image_file)
    img.width = 60
    img.height = 60
    return img

def writeToExcel(match, teamAinfo, teamAmainPerson, teamBinfo, teamBmainPerson, contrast):
    wb = load_workbook('template.xlsx')
    ws = wb.active
    ws['e3'] = teamAinfo['team_name']
    ws['f3'] = teamBinfo['team_name']

    imgA = downloadImg(match['team_A_logo'])
    imgB = downloadImg(match['team_B_logo'])
    ws.add_image(imgA, 'e4')
    ws.add_image(imgB, 'f4')
    

    start_time = utcstr2localstr(match['start_play'])
    date = start_time.split(' ')[0]
    ws['e6'] = date.replace('-', '月') + '日'
    time = start_time.split(' ')[1]
    if time[0] == '0':
        time = time[1:]
    ws['e5'] = time

    teamArank = teamAinfo['description'][0]
    
    teamBrank = teamBinfo['description'][0]
    ws['e9'] = teamArank['key'].replace('排名', ' ') + teamArank['value']
    ws['f9'] = teamBrank['key'].replace('排名', ' ') + teamBrank['value']


    comprehensive = contrast['comprehensive']
    ws['e10'] = comprehensive['team_A_score'].replace('%', '')
    ws['f10'] = comprehensive['team_B_score'].replace('%', '')

    prices = getProperty(comprehensive['data'], 'title', '身价')
    if prices:
        price_a = prices['team_A']['match_info']
        price_b = prices['team_B']['match_info']
        if '元' not in price_a:
            price_a += '元'
        if '元' not in price_b:
            price_b += '元'
        ws['e11'] = price_a
        ws['f11'] = price_b

    matchlast10 = getProperty(comprehensive['data'], 'title', '近10场战绩')
    if matchlast10:
        ws['e12'] = matchlast10['team_A']['match_info']
        ws['f12'] = matchlast10['team_B']['match_info']
    
    historyRecord = getProperty(comprehensive['data'], 'title', '近6场交锋')
    if historyRecord:
        ws['e18'] = teamAinfo['team_name'] + historyRecord['team_A']['match_info']

    scoreavg = getProperty(comprehensive['data'], 'title', '场均进球')
    if scoreavg:
        scoreavg_a = scoreavg['team_A']['match_info']
        scoreavg_b = scoreavg['team_B']['match_info']
        if '球' in scoreavg_a:
            scoreavg_a = scoreavg_a.replace('球', '')
        if '球' in scoreavg_b:
            scoreavg_b = scoreavg_b.replace('球', '')
        ws['e14'] = scoreavg_a
        ws['f14'] = scoreavg_b

    lostscoreavg = getProperty(comprehensive['data'], 'title', '场均失球')
    if lostscoreavg:
        lostscoreavg_a = lostscoreavg['team_A']['match_info']
        lostscoreavg_b = lostscoreavg['team_B']['match_info']
        if '球' in lostscoreavg_a:
            lostscoreavg_a = lostscoreavg_a.replace('球', '')
        if '球' in lostscoreavg_b:
            lostscoreavg_b = lostscoreavg_b.replace('球', '')
        ws['e15'] = lostscoreavg_a
        ws['f15'] = lostscoreavg_b

    control = contrast['control']['data']
    attacks = getProperty(control, 'title', '进攻次数')
    if attacks:
        attack_a = attacks['team_A']['match_info']
        attack_b = attacks['team_B']['match_info']
        if '次' in attack_a:
            attack_a = attack_a.replace('次', '')
        if '次' in attack_b:
            attack_b = attack_b.replace('次', '')
        ws['e16'] = attack_a
        ws['f16'] = attack_b

    ballcontrol = getProperty(control, 'title', '控球率')
    if ballcontrol:
        ws['e17'] = ballcontrol['team_A']['match_info']
        ws['f17'] = ballcontrol['team_B']['match_info']

    teamAscorePerson = getProperty(teamAmainPerson['person'], 'type', '进球')
    teamBscorePerson = getProperty(teamBmainPerson['person'], 'type', '进球')
    format = '%s %d球'
    if teamAscorePerson:
        ws['e21'] = format % (teamAscorePerson['person']['name'], teamAscorePerson['number'])
    if teamBscorePerson:
        ws['f21'] = format % (teamBscorePerson['person']['name'], teamBscorePerson['number'])


    format = '%s %d次'
    teamAAssistPerson = getProperty(teamAmainPerson['person'], 'type', '助攻')
    teamBAssistPerson = getProperty(teamBmainPerson['person'], 'type', '助攻')
    if teamAAssistPerson:
        ws['e22'] = format % (teamAAssistPerson['person']['name'], teamAAssistPerson['number'])
    if teamBAssistPerson:
        ws['f22'] = format % (teamBAssistPerson['person']['name'], teamBAssistPerson['number'])

    teamAMainPassPerson = getProperty(teamAmainPerson['person'], 'type', '关键传球')
    teamBMainPassPerson = getProperty(teamBmainPerson['person'], 'type', '关键传球')
    if teamAMainPassPerson:
        ws['e23'] = format % (teamAMainPassPerson['person']['name'], teamAMainPassPerson['number'])
    if teamBMainPassPerson:
        ws['f23'] = format % (teamBMainPassPerson['person']['name'], teamBMainPassPerson['number'])

    distfilename = 'static/%svs%s.xlsx' % (teamAinfo['team_name'], teamBinfo['team_name'])
    wb.save(distfilename)
    return distfilename.split('/')[1]